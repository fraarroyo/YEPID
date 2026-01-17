from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
import qrcode
from io import BytesIO
import base64
from PIL import Image
import os
import json
import uuid
from datetime import datetime, timedelta
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import sqlite3

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Change this to a secure random key in production

# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() in ['true', 'on', '1']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')  # Your email
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')  # Your email password or app password
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', app.config['MAIL_USERNAME'])

mail = Mail(app)

# Admin credentials (in production, store in database)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD_HASH = generate_password_hash("admin123")  # Default password: admin123

# Directory for storing generated QR codes and user data
QR_STORAGE_DIR = "static/qr_codes"
DATABASE = "yep_id.db"
os.makedirs(QR_STORAGE_DIR, exist_ok=True)

# SQLite Database Functions
def get_db():
    """Get database connection"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize database with tables"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            id TEXT,
            name TEXT NOT NULL,
            street TEXT,
            zone TEXT,
            sex TEXT,
            birthdate TEXT,
            email TEXT NOT NULL UNIQUE,
            phone TEXT,
            civil_status TEXT,
            youth_age_group TEXT,
            youth_classification TEXT,
            specific_needs_type TEXT,
            educational_background TEXT,
            educational_background_other TEXT,
            work_status TEXT,
            work_status_other TEXT,
            sk_voter_registered TEXT,
            sk_voted_last_election TEXT,
            national_voter_registered TEXT,
            attended_kk_assembly TEXT,
            kk_assembly_times TEXT,
            kk_assembly_no_reason TEXT,
            registration_date TEXT NOT NULL
        )
    ''')
    
    # Add new columns to existing table if they don't exist (for migration)
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN street TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN zone TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN sex TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN birthdate TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN civil_status TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN youth_age_group TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN youth_classification TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN specific_needs_type TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN educational_background TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN educational_background_other TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN work_status TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN work_status_other TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN sk_voter_registered TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN sk_voted_last_election TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN national_voter_registered TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN attended_kk_assembly TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN kk_assembly_times TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN kk_assembly_no_reason TEXT')
    except sqlite3.OperationalError:
        pass
    
    # Events table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS events (
            event_id TEXT PRIMARY KEY,
            event_name TEXT NOT NULL,
            event_year TEXT NOT NULL,
            event_description TEXT,
            event_date TEXT,
            event_time TEXT,
            event_points INTEGER DEFAULT 0,
            event_category TEXT,
            event_capacity INTEGER,
            reminder_sent INTEGER DEFAULT 0,
            created_date TEXT NOT NULL
        )
    ''')
    
    # Add new columns to events table if they don't exist
    try:
        cursor.execute('ALTER TABLE events ADD COLUMN event_category TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE events ADD COLUMN event_capacity INTEGER')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE events ADD COLUMN reminder_sent INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass
    
    # Attendance table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            attendance_id TEXT PRIMARY KEY,
            event_id TEXT NOT NULL,
            user_id TEXT NOT NULL,
            event_year TEXT,
            points_earned INTEGER DEFAULT 0,
            attendance_date TEXT NOT NULL,
            scan_time TEXT,
            FOREIGN KEY (event_id) REFERENCES events(event_id),
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    ''')
    
    # Add scan_time column if it doesn't exist
    try:
        cursor.execute('ALTER TABLE attendance ADD COLUMN scan_time TEXT')
    except sqlite3.OperationalError:
        pass
    
    # Notifications table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            notification_id TEXT PRIMARY KEY,
            user_id TEXT,
            event_id TEXT,
            notification_type TEXT NOT NULL,
            notification_title TEXT NOT NULL,
            notification_message TEXT NOT NULL,
            sent_date TEXT NOT NULL,
            read_status INTEGER DEFAULT 0,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            FOREIGN KEY (event_id) REFERENCES events(event_id)
        )
    ''')
    
    # Create indexes for better performance
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_event ON attendance(event_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_user ON attendance(user_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_year ON attendance(event_year)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)')
    
    conn.commit()
    conn.close()

def migrate_json_to_db():
    """Migrate existing JSON data to SQLite database"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if database is already populated
    cursor.execute('SELECT COUNT(*) FROM users')
    if cursor.fetchone()[0] > 0:
        conn.close()
        return  # Already migrated
    
    # Migrate users
    if os.path.exists("users_data.json"):
        try:
            with open("users_data.json", 'r') as f:
                users = json.load(f)
                for user in users:
                    cursor.execute('''
                        INSERT OR IGNORE INTO users 
                        (user_id, id, name, street, zone, sex, birthdate, email, phone, 
                         civil_status, youth_age_group, youth_classification, specific_needs_type,
                         educational_background, educational_background_other, work_status, work_status_other,
                         sk_voter_registered, sk_voted_last_election, national_voter_registered, attended_kk_assembly,
                         kk_assembly_times, kk_assembly_no_reason, registration_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        user.get('user_id'),
                        user.get('id', ''),
                        user.get('name', ''),
                        user.get('street', ''),
                        user.get('zone', ''),
                        user.get('sex', ''),
                        user.get('birthdate', ''),
                        user.get('email', ''),
                        user.get('phone', ''),
                        user.get('civil_status', ''),
                        user.get('youth_age_group', ''),
                        user.get('youth_classification', ''),
                        user.get('specific_needs_type', ''),
                        user.get('educational_background', ''),
                        user.get('educational_background_other', ''),
                        user.get('work_status', ''),
                        user.get('work_status_other', ''),
                        user.get('sk_voter_registered', ''),
                        user.get('sk_voted_last_election', ''),
                        user.get('national_voter_registered', ''),
                        user.get('attended_kk_assembly', ''),
                        user.get('kk_assembly_times', ''),
                        user.get('kk_assembly_no_reason', ''),
                        user.get('registration_date', datetime.now().isoformat())
                    ))
        except Exception as e:
            print(f"Error migrating users: {e}")
    
    # Migrate events
    if os.path.exists("events_data.json"):
        try:
            with open("events_data.json", 'r') as f:
                events = json.load(f)
                for event in events:
                    cursor.execute('''
                        INSERT OR IGNORE INTO events 
                        (event_id, event_name, event_year, event_description, event_date, event_time, event_points, created_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        event.get('event_id'),
                        event.get('event_name', ''),
                        event.get('event_year', ''),
                        event.get('event_description', ''),
                        event.get('event_date', ''),
                        event.get('event_time', ''),
                        event.get('event_points', 0),
                        event.get('created_date', datetime.now().isoformat())
                    ))
        except Exception as e:
            print(f"Error migrating events: {e}")
    
    # Migrate attendance
    if os.path.exists("attendance_data.json"):
        try:
            with open("attendance_data.json", 'r') as f:
                attendance = json.load(f)
                for record in attendance:
                    cursor.execute('''
                        INSERT OR IGNORE INTO attendance 
                        (attendance_id, event_id, user_id, event_year, points_earned, attendance_date)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        record.get('attendance_id'),
                        record.get('event_id'),
                        record.get('user_id'),
                        record.get('event_year', ''),
                        record.get('points_earned', 0),
                        record.get('attendance_date', datetime.now().isoformat())
                    ))
        except Exception as e:
            print(f"Error migrating attendance: {e}")
    
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()
migrate_json_to_db()

def load_users():
    """Load users from database"""
    conn = get_db()
    cursor = conn.cursor()
    # Order by ID numbers in ascending order (Youth001, Youth002, etc.)
    cursor.execute('''
        SELECT * FROM users 
        ORDER BY 
            CASE 
                WHEN id LIKE 'Youth%' THEN CAST(SUBSTR(id, 6) AS INTEGER)
                ELSE 999999
            END ASC,
            registration_date ASC
    ''')
    rows = cursor.fetchall()
    users = [dict(row) for row in rows]
    conn.close()
    return users

def save_user(user_data):
    """Save user to database"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO users (user_id, id, name, street, zone, sex, birthdate, email, phone, 
                          civil_status, youth_age_group, youth_classification, specific_needs_type,
                          educational_background, educational_background_other, work_status, work_status_other,
                          sk_voter_registered, sk_voted_last_election, national_voter_registered, attended_kk_assembly,
                          kk_assembly_times, kk_assembly_no_reason, registration_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        user_data.get('user_id'),
        user_data.get('id', ''),
        user_data.get('name', ''),
        user_data.get('street', ''),
        user_data.get('zone', ''),
        user_data.get('sex', ''),
        user_data.get('birthdate', ''),
        user_data.get('email', ''),
        user_data.get('phone', ''),
        user_data.get('civil_status', ''),
        user_data.get('youth_age_group', ''),
        user_data.get('youth_classification', ''),
        user_data.get('specific_needs_type', ''),
        user_data.get('educational_background', ''),
        user_data.get('educational_background_other', ''),
        user_data.get('work_status', ''),
        user_data.get('work_status_other', ''),
        user_data.get('sk_voter_registered', ''),
        user_data.get('sk_voted_last_election', ''),
        user_data.get('national_voter_registered', ''),
        user_data.get('attended_kk_assembly', ''),
        user_data.get('kk_assembly_times', ''),
        user_data.get('kk_assembly_no_reason', ''),
        user_data.get('registration_date', datetime.now().isoformat())
    ))
    conn.commit()
    conn.close()

def load_events():
    """Load events from database"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM events ORDER BY event_year DESC, created_date DESC')
    rows = cursor.fetchall()
    events = [dict(row) for row in rows]
    conn.close()
    return events

def save_event(event_data):
    """Save event to database"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO events 
        (event_id, event_name, event_year, event_description, event_date, event_time, event_points, event_category, event_capacity, reminder_sent, created_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        event_data.get('event_id'),
        event_data.get('event_name', ''),
        event_data.get('event_year', ''),
        event_data.get('event_description', ''),
        event_data.get('event_date', ''),
        event_data.get('event_time', ''),
        event_data.get('event_points', 0),
        event_data.get('event_category', ''),
        event_data.get('event_capacity'),
        event_data.get('reminder_sent', 0),
        event_data.get('created_date', datetime.now().isoformat())
    ))
    conn.commit()
    conn.close()

def update_events(events):
    """Update events in database (for deletion)"""
    # This function is mainly used for deleting events
    # Individual updates are handled by delete_event route
    pass

def load_attendance():
    """Load attendance records from database"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM attendance ORDER BY attendance_date DESC')
    rows = cursor.fetchall()
    attendance = [dict(row) for row in rows]
    conn.close()
    return attendance

def save_attendance(attendance_data):
    """Save attendance record to database"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO attendance 
        (attendance_id, event_id, user_id, event_year, points_earned, attendance_date, scan_time)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        attendance_data.get('attendance_id'),
        attendance_data.get('event_id'),
        attendance_data.get('user_id'),
        attendance_data.get('event_year', ''),
        attendance_data.get('points_earned', 0),
        attendance_data.get('attendance_date', datetime.now().isoformat()),
        attendance_data.get('scan_time', datetime.now().isoformat())
    ))
    conn.commit()
    conn.close()

# Analytics and Reporting Functions
def get_analytics_stats():
    """Get overall analytics statistics"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Total registrations
    cursor.execute('SELECT COUNT(*) FROM users')
    total_users = cursor.fetchone()[0]
    
    # Total events
    cursor.execute('SELECT COUNT(*) FROM events')
    total_events = cursor.fetchone()[0]
    
    # Active events (upcoming)
    today = datetime.now().date().isoformat()
    cursor.execute('SELECT COUNT(*) FROM events WHERE event_date >= ?', (today,))
    active_events = cursor.fetchone()[0]
    
    # Total attendance records
    cursor.execute('SELECT COUNT(*) FROM attendance')
    total_attendance = cursor.fetchone()[0]
    
    # Total points distributed
    cursor.execute('SELECT SUM(points_earned) FROM attendance')
    result = cursor.fetchone()[0]
    total_points = result if result else 0
    
    # Average attendance per event
    cursor.execute('''
        SELECT AVG(attendance_count) FROM (
            SELECT COUNT(*) as attendance_count 
            FROM attendance 
            GROUP BY event_id
        )
    ''')
    result = cursor.fetchone()[0]
    avg_attendance = round(result, 1) if result else 0
    
    # Recent registrations (last 7 days)
    seven_days_ago = (datetime.now() - timedelta(days=7)).isoformat()
    cursor.execute('SELECT COUNT(*) FROM users WHERE registration_date >= ?', (seven_days_ago,))
    recent_registrations = cursor.fetchone()[0]
    
    conn.close()
    
    return {
        'total_users': total_users,
        'total_events': total_events,
        'active_events': active_events,
        'total_attendance': total_attendance,
        'total_points': total_points,
        'avg_attendance': avg_attendance,
        'recent_registrations': recent_registrations
    }

def get_demographic_stats():
    """Get demographic breakdown statistics"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Age group breakdown
    cursor.execute('SELECT youth_age_group, COUNT(*) as count FROM users WHERE youth_age_group IS NOT NULL AND youth_age_group != "" GROUP BY youth_age_group')
    age_groups = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Zone breakdown
    cursor.execute('SELECT zone, COUNT(*) as count FROM users WHERE zone IS NOT NULL AND zone != "" GROUP BY zone')
    zones = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Classification breakdown
    cursor.execute('SELECT youth_classification, COUNT(*) as count FROM users WHERE youth_classification IS NOT NULL AND youth_classification != "" GROUP BY youth_classification')
    classifications = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Sex breakdown
    cursor.execute('SELECT sex, COUNT(*) as count FROM users WHERE sex IS NOT NULL AND sex != "" GROUP BY sex')
    sex_breakdown = {row[0]: row[1] for row in cursor.fetchall()}
    
    conn.close()
    
    return {
        'age_groups': age_groups,
        'zones': zones,
        'classifications': classifications,
        'sex_breakdown': sex_breakdown
    }

def get_event_analytics():
    """Get event analytics and trends"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Events with attendance counts
    cursor.execute('''
        SELECT e.event_id, e.event_name, e.event_date, e.event_points,
               COUNT(a.attendance_id) as attendance_count,
               e.event_capacity
        FROM events e
        LEFT JOIN attendance a ON e.event_id = a.event_id
        GROUP BY e.event_id
        ORDER BY e.event_date DESC
    ''')
    
    events_data = []
    for row in cursor.fetchall():
        events_data.append({
            'event_id': row[0],
            'event_name': row[1],
            'event_date': row[2],
            'event_points': row[3],
            'attendance_count': row[4],
            'event_capacity': row[5],
            'attendance_rate': round((row[4] / row[5] * 100), 1) if row[5] and row[5] > 0 else 0
        })
    
    # Monthly attendance trends
    cursor.execute('''
        SELECT strftime('%Y-%m', attendance_date) as month, COUNT(*) as count
        FROM attendance
        GROUP BY month
        ORDER BY month DESC
        LIMIT 12
    ''')
    monthly_trends = {row[0]: row[1] for row in cursor.fetchall()}
    
    conn.close()
    
    return {
        'events_data': events_data,
        'monthly_trends': monthly_trends
    }

# Email Notification Functions
def send_email_notification(recipient_email, subject, message_body, html_body=None):
    """Send email notification"""
    try:
        msg = Message(
            subject=subject,
            recipients=[recipient_email],
            body=message_body,
            html=html_body
        )
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def send_attendance_confirmation(user_email, user_name, event_name, points_earned):
    """Send attendance confirmation email"""
    subject = f"Attendance Confirmed - {event_name}"
    message_body = f"""
Dear {user_name},

Your attendance for the event "{event_name}" has been successfully recorded.

Points Earned: {points_earned}

Thank you for your participation!

Best regards,
SAN AGUSTIN YEP ID System
    """
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h2 style="color: #002e6a;">Attendance Confirmed</h2>
        <p>Dear {user_name},</p>
        <p>Your attendance for the event <strong>"{event_name}"</strong> has been successfully recorded.</p>
        <p><strong>Points Earned:</strong> {points_earned}</p>
        <p>Thank you for your participation!</p>
        <p>Best regards,<br>SAN AGUSTIN YEP ID System</p>
    </body>
    </html>
    """
    return send_email_notification(user_email, subject, message_body, html_body)

def send_event_reminder(user_email, user_name, event_name, event_date, event_time):
    """Send event reminder email"""
    subject = f"Reminder: {event_name}"
    message_body = f"""
Dear {user_name},

This is a reminder that you have an upcoming event:

Event: {event_name}
Date: {event_date}
Time: {event_time}

We look forward to seeing you there!

Best regards,
SAN AGUSTIN YEP ID System
    """
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h2 style="color: #002e6a;">Event Reminder</h2>
        <p>Dear {user_name},</p>
        <p>This is a reminder that you have an upcoming event:</p>
        <ul>
            <li><strong>Event:</strong> {event_name}</li>
            <li><strong>Date:</strong> {event_date}</li>
            <li><strong>Time:</strong> {event_time}</li>
        </ul>
        <p>We look forward to seeing you there!</p>
        <p>Best regards,<br>SAN AGUSTIN YEP ID System</p>
    </body>
    </html>
    """
    return send_email_notification(user_email, subject, message_body, html_body)

def send_points_update(user_email, user_name, total_points, events_attended):
    """Send points update email"""
    subject = "Your Points Update - SAN AGUSTIN YEP ID"
    message_body = f"""
Dear {user_name},

Your current points status:

Total Points: {total_points}
Events Attended: {events_attended}

Keep up the great participation!

Best regards,
SAN AGUSTIN YEP ID System
    """
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h2 style="color: #002e6a;">Points Update</h2>
        <p>Dear {user_name},</p>
        <p>Your current points status:</p>
        <ul>
            <li><strong>Total Points:</strong> {total_points}</li>
            <li><strong>Events Attended:</strong> {events_attended}</li>
        </ul>
        <p>Keep up the great participation!</p>
        <p>Best regards,<br>SAN AGUSTIN YEP ID System</p>
    </body>
    </html>
    """
    return send_email_notification(user_email, subject, message_body, html_body)

def generate_user_qr_code(user_data, save_to_disk=True):
    """Generate QR code for user with their registration data"""
    # Create a unique identifier for the user
    qr_data = json.dumps({
        'user_id': user_data['user_id'],
        'name': user_data['name'],
        'email': user_data['email'],
        'registration_date': user_data['registration_date']
    })
    
    # Generate QR code with better error correction for easier scanning
    qr = qrcode.QRCode(
        version=None,  # Auto-detect version
        error_correction=qrcode.constants.ERROR_CORRECT_M,  # Medium error correction
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    # Create image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to BytesIO
    img_buffer = BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    # Save to disk if requested
    if save_to_disk:
        qr_filename = f"{user_data['user_id']}.png"
        qr_filepath = os.path.join(QR_STORAGE_DIR, qr_filename)
        img.save(qr_filepath, format='PNG')
    
    return img_buffer

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    """Home page - redirects to consent"""
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('consent'))

@app.route('/consent', methods=['GET', 'POST'])
def consent():
    """Informed consent page"""
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        consent_value = request.form.get('consent', '').strip()
        
        if consent_value == 'agree':
            # Set consent in session
            session['consent_given'] = True
            flash('Thank you for your consent. You may now proceed with registration.', 'success')
            return redirect(url_for('register'))
        elif consent_value == 'disagree':
            flash('You have chosen not to participate. Registration cannot proceed without your consent.', 'error')
            return render_template('consent.html')
    
    return render_template('consent.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration page"""
    # Check if user has given consent
    if 'logged_in' not in session and not session.get('consent_given'):
        flash('Please provide your consent before proceeding with registration.', 'warning')
        return redirect(url_for('consent'))
    
    if request.method == 'POST':
        # Get all form fields
        name = request.form.get('name', '').strip()
        street = request.form.get('street', '').strip()
        zone = request.form.get('zone', '').strip()
        sex = request.form.get('sex', '').strip()
        birthdate = request.form.get('birthdate', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        civil_status = request.form.get('civil_status', '').strip()
        youth_age_group = request.form.get('youth_age_group', '').strip()
        youth_classification = request.form.get('youth_classification', '').strip()
        specific_needs_type = request.form.get('specific_needs_type', '').strip()
        educational_background = request.form.get('educational_background', '').strip()
        educational_background_other = request.form.get('educational_background_other', '').strip()
        work_status = request.form.get('work_status', '').strip()
        work_status_other = request.form.get('work_status_other', '').strip()
        sk_voter_registered = request.form.get('sk_voter_registered', '').strip()
        sk_voted_last_election = request.form.get('sk_voted_last_election', '').strip()
        national_voter_registered = request.form.get('national_voter_registered', '').strip()
        attended_kk_assembly = request.form.get('attended_kk_assembly', '').strip()
        kk_assembly_times = request.form.get('kk_assembly_times', '').strip()
        kk_assembly_no_reason = request.form.get('kk_assembly_no_reason', '').strip()
        
        # Validation
        if not name or not email:
            flash('Please fill in all required fields (Full Name and Email).', 'error')
            return render_template('register.html')
        
        # Check if email already exists
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT email FROM users WHERE LOWER(email) = LOWER(?)', (email,))
        if cursor.fetchone():
            conn.close()
            flash('This email is already registered.', 'error')
            return render_template('register.html')
        
        # Generate sequential ID based on registration order (first-come-first-serve)
        # Count total registered users to get the next sequential number
        cursor.execute('SELECT COUNT(*) FROM users')
        total_users = cursor.fetchone()[0]
        next_num = total_users + 1
        generated_id = f"Youth{next_num:03d}"
        conn.close()
        
        # Create user data
        user_id = str(uuid.uuid4())
        user_data = {
            'user_id': user_id,
            'id': generated_id,
            'name': name,
            'street': street,
            'zone': zone,
            'sex': sex,
            'birthdate': birthdate,
            'email': email,
            'phone': phone,
            'civil_status': civil_status,
            'youth_age_group': youth_age_group,
            'youth_classification': youth_classification,
            'specific_needs_type': specific_needs_type,
            'educational_background': educational_background,
            'educational_background_other': educational_background_other,
            'work_status': work_status,
            'work_status_other': work_status_other,
            'sk_voter_registered': sk_voter_registered,
            'sk_voted_last_election': sk_voted_last_election,
            'national_voter_registered': national_voter_registered,
            'attended_kk_assembly': attended_kk_assembly,
            'kk_assembly_times': kk_assembly_times,
            'kk_assembly_no_reason': kk_assembly_no_reason,
            'registration_date': datetime.now().isoformat()
        }
        
        # Generate QR code
        try:
            qr_buffer = generate_user_qr_code(user_data)
            qr_buffer.seek(0)
            
            # Save user data
            save_user(user_data)
            
            # Send email with QR code
            try:
                msg = Message(
                    subject='Your Registration QR Code - SAN AGUSTIN YEP ID',
                    recipients=[email],
                    html=f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                            <h2 style="color: #4a90e2;">Welcome, {name}!</h2>
                            <p>Thank you for registering with SAN AGUSTIN YEP ID system.</p>
                            <p>Your registration has been successful. Please find your unique QR code attached to this email.</p>
                            <p><strong>Registration Details:</strong></p>
                            <ul>
                                <li><strong>Name:</strong> {name}</li>
                                <li><strong>Email:</strong> {email}</li>
                                <li><strong>Phone:</strong> {phone if phone else 'Not provided'}</li>
                                <li><strong>Registration Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
                            </ul>
                            <p>Please keep this QR code safe. You can use it for identification and verification purposes.</p>
                            <p>If you have any questions, please contact our support team.</p>
                            <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
                            <p style="color: #666; font-size: 12px;">This is an automated message. Please do not reply to this email.</p>
                        </div>
                    </body>
                    </html>
                    """
                )
                msg.attach('qrcode.png', 'image/png', qr_buffer.read(), 'inline', headers=[['Content-ID', '<qrcode>']])
                
                mail.send(msg)
                flash('Registration successful! Your QR code has been sent to your email.', 'success')
            except Exception as e:
                # If email fails, still save the user but show warning
                flash(f'Registration successful, but email could not be sent: {str(e)}. Please contact support.', 'warning')
            
            # Clear consent session so user must consent again for next registration
            session.pop('consent_given', None)
            return redirect(url_for('registration_success', email=email))
            
        except Exception as e:
            flash(f'Error generating QR code: {str(e)}', 'error')
            return render_template('register.html')
    
    return render_template('register.html')

@app.route('/registration_success')
def registration_success():
    """Registration success page"""
    email = request.args.get('email', '')
    return render_template('registration_success.html', email=email)

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['logged_in'] = True
            session['username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout route"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard after login"""
    stats = get_analytics_stats()
    return render_template('dashboard.html', stats=stats)

@app.route('/registered_persons')
@login_required
def registered_persons():
    """View all registered persons - Admin only"""
    users = load_users()
    
    # Check which users have QR codes saved and generate missing ones
    # Also ensure all users have the id field for backward compatibility
    conn = get_db()
    cursor = conn.cursor()
    
    # First, assign IDs to users without IDs based on registration order
    cursor.execute('''
        SELECT user_id, registration_date 
        FROM users 
        WHERE id IS NULL OR id = "" OR id = "STU%" 
        ORDER BY registration_date ASC
    ''')
    users_without_id = cursor.fetchall()
    
    # Get the highest existing Youth ID number
    cursor.execute('SELECT id FROM users WHERE id LIKE "Youth%" ORDER BY CAST(SUBSTR(id, 6) AS INTEGER) DESC LIMIT 1')
    last_id_row = cursor.fetchone()
    if last_id_row and last_id_row[0]:
        last_id = last_id_row[0]
        if last_id.startswith('Youth') and last_id.replace('Youth', '').isdigit():
            next_id_num = int(last_id.replace('Youth', '')) + 1
        else:
            next_id_num = 1
    else:
        next_id_num = 1
    
    # Assign IDs sequentially to users without IDs, maintaining registration order
    for user_row in users_without_id:
        user_id_to_update = user_row[0]
        new_id = f"Youth{next_id_num:03d}"
        cursor.execute('UPDATE users SET id = ? WHERE user_id = ?', (new_id, user_id_to_update))
        next_id_num += 1
    
    # Reload users to get updated IDs, ordered by ID numbers
    cursor.execute('''
        SELECT * FROM users 
        ORDER BY 
            CASE 
                WHEN id LIKE 'Youth%' THEN CAST(SUBSTR(id, 6) AS INTEGER)
                ELSE 999999
            END ASC,
            registration_date ASC
    ''')
    rows = cursor.fetchall()
    users = [dict(row) for row in rows]
    
    for user in users:
        # Migrate old STU IDs to Youth format
        if user.get('id', '').startswith('STU'):
            old_num = user['id'].replace('STU', '')
            if old_num.isdigit():
                new_id = f"Youth{old_num}"
                cursor.execute('UPDATE users SET id = ? WHERE user_id = ?', (new_id, user['user_id']))
                user['id'] = new_id
        
        qr_filename = f"{user['user_id']}.png"
        qr_filepath = os.path.join(QR_STORAGE_DIR, qr_filename)
        if not os.path.exists(qr_filepath):
            # Generate QR code for users who don't have one
            try:
                generate_user_qr_code(user, save_to_disk=True)
            except Exception as e:
                print(f"Error generating QR code for user {user['user_id']}: {str(e)}")
    
    conn.commit()
    conn.close()
    
    # Reload users to get updated data
    users = load_users()
    
    return render_template('registered_persons.html', users=users)

@app.route('/generate_user_qr/<user_id>')
@login_required
def generate_user_qr(user_id):
    """Generate and download QR code for a specific registered person"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE user_id = ?', (user_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        flash('User not found.', 'error')
        return redirect(url_for('registered_persons'))
    
    user = dict(row)
    
    # Generate and save QR code
    qr_buffer = generate_user_qr_code(user, save_to_disk=True)
    qr_buffer.seek(0)
    
    return send_file(qr_buffer, mimetype='image/png', as_attachment=True, 
                    download_name=f"qr_{user['name'].replace(' ', '_')}_{user_id[:8]}.png")

@app.route('/view_user_qr/<user_id>')
@login_required
def view_user_qr(user_id):
    """View QR code for a specific registered person"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE user_id = ?', (user_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        flash('User not found.', 'error')
        return redirect(url_for('registered_persons'))
    
    user = dict(row)
    
    qr_filename = f"{user_id}.png"
    qr_filepath = os.path.join(QR_STORAGE_DIR, qr_filename)
    
    # Generate QR code if it doesn't exist
    if not os.path.exists(qr_filepath):
        generate_user_qr_code(user, save_to_disk=True)
    
    if os.path.exists(qr_filepath):
        return send_file(qr_filepath, mimetype='image/png')
    else:
        flash('QR code could not be generated.', 'error')
        return redirect(url_for('registered_persons'))

@app.route('/events', methods=['GET', 'POST'])
@login_required
def events():
    """Manage events - view all events and create new ones"""
    if request.method == 'POST':
        event_name = request.form.get('event_name', '').strip()
        event_description = request.form.get('event_description', '').strip()
        event_date = request.form.get('event_date', '').strip()
        event_time = request.form.get('event_time', '').strip()
        event_points = request.form.get('event_points', '0').strip()
        event_category = request.form.get('event_category', '').strip()
        event_capacity = request.form.get('event_capacity', '').strip()
        
        if not event_name or not event_date:
            flash('Please fill in event name and date.', 'error')
            return redirect(url_for('events'))
        
        try:
            event_points = int(event_points) if event_points else 0
        except ValueError:
            event_points = 0
        
        try:
            event_capacity = int(event_capacity) if event_capacity else None
        except ValueError:
            event_capacity = None
        
        # Extract year from date (format: YYYY-MM-DD)
        event_year = event_date.split('-')[0] if event_date else ''
        
        event_id = str(uuid.uuid4())
        event_data = {
            'event_id': event_id,
            'event_name': event_name,
            'event_year': event_year,
            'event_description': event_description,
            'event_date': event_date,
            'event_time': event_time,
            'event_points': event_points,
            'event_category': event_category,
            'event_capacity': event_capacity,
            'created_date': datetime.now().isoformat()
        }
        
        save_event(event_data)
        flash('Event created successfully!', 'success')
        return redirect(url_for('events'))
    
    events_list = load_events()  # Already sorted by year and date in SQL query
    return render_template('events.html', events=events_list)

@app.route('/events/<event_id>')
@login_required
def event_detail(event_id):
    """View event details and manage attendance"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get event
    cursor.execute('SELECT * FROM events WHERE event_id = ?', (event_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        flash('Event not found.', 'error')
        return redirect(url_for('events'))
    event = dict(row)
    
    # Load attendance for this event with user details
    cursor.execute('''
        SELECT a.*, u.name as user_name, u.email as user_email
        FROM attendance a
        LEFT JOIN users u ON a.user_id = u.user_id
        WHERE a.event_id = ?
        ORDER BY a.attendance_date DESC
    ''', (event_id,))
    rows = cursor.fetchall()
    event_attendance = [dict(row) for row in rows]
    
    # Ensure points_earned is set (for backward compatibility)
    for record in event_attendance:
        if 'points_earned' not in record or record.get('points_earned') is None:
            record['points_earned'] = event.get('event_points', 0)
    
    conn.close()
    return render_template('event_detail.html', event=event, attendance=event_attendance)

@app.route('/events/<event_id>/scan', methods=['GET', 'POST'])
@login_required
def scan_event_attendance(event_id):
    """Scan QR code for event attendance"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM events WHERE event_id = ?', (event_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        flash('Event not found.', 'error')
        return redirect(url_for('events'))
    
    event = dict(row)
    return render_template('scan_event_attendance.html', event=event)

@app.route('/api/scan/attendance/<event_id>', methods=['POST'])
@login_required
def process_attendance_scan(event_id):
    """Process scanned QR code for event attendance"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get event
    cursor.execute('SELECT * FROM events WHERE event_id = ?', (event_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return json.dumps({'success': False, 'error': 'Event not found'}), 404
    event = dict(row)
    
    data = request.get_json()
    scanned_data = data.get('qr_data', '')
    
    if not scanned_data:
        conn.close()
        return json.dumps({'success': False, 'error': 'No QR code data provided'}), 400
    
    try:
        # Parse QR code data (should be JSON)
        qr_user_data = json.loads(scanned_data)
        user_id = qr_user_data.get('user_id')
        
        if user_id:
            # Check if user exists
            cursor.execute('SELECT * FROM users WHERE user_id = ?', (user_id,))
            user_row = cursor.fetchone()
            
            if user_row:
                user = dict(user_row)
                
                # Check if already attended
                cursor.execute('SELECT * FROM attendance WHERE event_id = ? AND user_id = ?', (event_id, user_id))
                existing_attendance = cursor.fetchone()
                
                if not existing_attendance:
                    # Get event points
                    event_points = event.get('event_points', 0)
                    
                    # Record attendance
                    attendance_id = str(uuid.uuid4())
                    scan_time = datetime.now().isoformat()
                    cursor.execute('''
                        INSERT INTO attendance 
                        (attendance_id, event_id, user_id, event_year, points_earned, attendance_date, scan_time)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        attendance_id,
                        event_id,
                        user_id,
                        event.get('event_year', ''),
                        event_points,
                        datetime.now().isoformat(),
                        scan_time
                    ))
                    conn.commit()
                    conn.close()
                    
                    # Send attendance confirmation email
                    try:
                        send_attendance_confirmation(
                            user.get('email', ''),
                            user.get('name', 'Unknown'),
                            event.get('event_name', 'Event'),
                            event_points
                        )
                    except Exception as e:
                        print(f"Error sending attendance confirmation email: {e}")
                    
                    return json.dumps({
                        'success': True, 
                        'message': f'Attendance recorded for {user.get("name", "Unknown")}!',
                        'user_name': user.get('name', 'Unknown')
                    })
                else:
                    conn.close()
                    return json.dumps({
                        'success': False, 
                        'error': f'{user.get("name", "Unknown")} has already been marked as attended for this event.',
                        'already_attended': True
                    }), 400
            else:
                conn.close()
                return json.dumps({'success': False, 'error': 'User not found in the system.'}), 400
        else:
            conn.close()
            return json.dumps({'success': False, 'error': 'Invalid QR code format. User ID not found.'}), 400
    except json.JSONDecodeError:
        conn.close()
        return json.dumps({'success': False, 'error': 'Invalid QR code format. Could not parse user data.'}), 400
    except Exception as e:
        conn.close()
        return json.dumps({'success': False, 'error': f'Error processing attendance: {str(e)}'}), 500

@app.route('/events/<event_id>/delete', methods=['POST'])
@login_required
def delete_event(event_id):
    """Delete an event"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if event exists
    cursor.execute('SELECT * FROM events WHERE event_id = ?', (event_id,))
    if not cursor.fetchone():
        conn.close()
        flash('Event not found.', 'error')
        return redirect(url_for('events'))
    
    # Delete attendance records for this event
    cursor.execute('DELETE FROM attendance WHERE event_id = ?', (event_id,))
    
    # Delete event
    cursor.execute('DELETE FROM events WHERE event_id = ?', (event_id,))
    
    conn.commit()
    conn.close()
    
    flash('Event deleted successfully!', 'success')
    return redirect(url_for('events'))

@app.route('/events/<event_id>/export')
@login_required
def export_attendance(event_id):
    """Export attendance records to Excel file"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get event
    cursor.execute('SELECT * FROM events WHERE event_id = ?', (event_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        flash('Event not found.', 'error')
        return redirect(url_for('events'))
    event = dict(row)
    
    # Load attendance for this event with user details
    cursor.execute('''
        SELECT a.*, u.name, u.email, u.id, u.phone
        FROM attendance a
        LEFT JOIN users u ON a.user_id = u.user_id
        WHERE a.event_id = ?
        ORDER BY a.attendance_date ASC, a.attendance_date ASC
    ''', (event_id,))
    rows = cursor.fetchall()
    conn.close()
    
    attendance_data = []
    for row in rows:
        record = dict(row)
        # Parse attendance date
        att_date_str = record.get('attendance_date', '')
        if 'T' in att_date_str:
            att_date = att_date_str.split('T')[0]
            att_time = att_date_str.split('T')[1].split('.')[0]
        else:
            att_date = att_date_str
            att_time = 'N/A'
        
        attendance_data.append({
            'name': record.get('name', 'Unknown'),
            'email': record.get('email', 'Unknown'),
            'id': record.get('id', 'N/A'),
            'phone': record.get('phone', 'N/A'),
            'attendance_date': att_date,
            'attendance_time': att_time,
            'points_earned': record.get('points_earned', event.get('event_points', 0))
        })
    
    # Sort by attendance date and time
    attendance_data.sort(key=lambda x: (x['attendance_date'], x['attendance_time']))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Header row
    headers = ['#', 'Name', 'ID', 'Email', 'Phone', 'Points Earned', 'Attendance Date', 'Attendance Time']
    header_fill = PatternFill(start_color="002e6a", end_color="002e6a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, record in enumerate(attendance_data, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)  # Serial number
        ws.cell(row=row_num, column=2, value=record['name'])
        ws.cell(row=row_num, column=3, value=record['id'])
        ws.cell(row=row_num, column=4, value=record['email'])
        ws.cell(row=row_num, column=5, value=record['phone'])
        ws.cell(row=row_num, column=6, value=record['points_earned'])
        ws.cell(row=row_num, column=7, value=record['attendance_date'])
        ws.cell(row=row_num, column=8, value=record['attendance_time'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    # Add event information sheet
    ws_info = wb.create_sheet("Event Information")
    ws_info.append(['Event Name', event.get('event_name', 'N/A')])
    ws_info.append(['Year', event.get('event_year', 'N/A')])
    ws_info.append(['Event Date', event.get('event_date', 'N/A')])
    ws_info.append(['Event Time', event.get('event_time', 'N/A')])
    ws_info.append(['Description', event.get('event_description', 'N/A')])
    ws_info.append(['Total Attendance', len(attendance_data)])
    ws_info.append(['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    # Adjust info sheet column widths
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 40
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename
    event_name_safe = "".join(c for c in event.get('event_name', 'Event') if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"Attendance_{event_name_safe}_{event.get('event_year', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/leaderboard')
@login_required
def leaderboard():
    """View points leaderboard"""
    year = request.args.get('year', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Build query with optional year filter
    if year:
        query = '''
            SELECT 
                u.user_id,
                u.name,
                u.email,
                u.id,
                COALESCE(SUM(a.points_earned), 0) as total_points,
                COUNT(a.attendance_id) as events_attended
            FROM users u
            LEFT JOIN attendance a ON u.user_id = a.user_id AND a.event_year = ?
            GROUP BY u.user_id, u.name, u.email, u.id
            HAVING total_points > 0
            ORDER BY total_points DESC, events_attended DESC
        '''
        cursor.execute(query, (year,))
    else:
        query = '''
            SELECT 
                u.user_id,
                u.name,
                u.email,
                u.id,
                COALESCE(SUM(a.points_earned), 0) as total_points,
                COUNT(a.attendance_id) as events_attended
            FROM users u
            LEFT JOIN attendance a ON u.user_id = a.user_id
            GROUP BY u.user_id, u.name, u.email, u.id
            HAVING total_points > 0
            ORDER BY total_points DESC, events_attended DESC
        '''
        cursor.execute(query)
    
    rows = cursor.fetchall()
    leaderboard_data = [dict(row) for row in rows]
    
    # Get available years for filter
    cursor.execute('SELECT DISTINCT event_year FROM attendance WHERE event_year IS NOT NULL AND event_year != "" ORDER BY event_year DESC')
    year_rows = cursor.fetchall()
    available_years = [row[0] for row in year_rows]
    
    conn.close()
    
    return render_template('leaderboard.html', 
                         leaderboard=leaderboard_data, 
                         selected_year=year,
                         available_years=available_years)

@app.route('/generate', methods=['GET', 'POST'])
@login_required
def generate_qr():
    """Generate QR code page"""
    qr_image = None
    qr_data = None
    
    if request.method == 'POST':
        data = request.form.get('qr_data', '').strip()
        
        if data:
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(data)
            qr.make(fit=True)
            
            # Create image
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Convert to base64 for display
            img_buffer = BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
            qr_image = f"data:image/png;base64,{img_base64}"
            qr_data = data
            
            flash('QR code generated successfully!', 'success')
        else:
            flash('Please enter data to generate QR code.', 'error')
    
    return render_template('generate.html', qr_image=qr_image, qr_data=qr_data)

@app.route('/scan', methods=['GET', 'POST'])
@login_required
def scan_qr():
    """QR code scanning page"""
    return render_template('scan.html')

@app.route('/api/scan/process', methods=['POST'])
@login_required
def process_scan():
    """Process scanned QR code data"""
    data = request.get_json()
    scanned_data = data.get('qr_data', '')
    
    if scanned_data:
        return json.dumps({'success': True, 'data': scanned_data})
    else:
        return json.dumps({'success': False, 'error': 'No QR code data provided'}), 400

@app.route('/download_qr')
@login_required
def download_qr():
    """Download QR code as image"""
    data = request.args.get('data', '')
    
    if data:
        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to BytesIO
        img_buffer = BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        return send_file(img_buffer, mimetype='image/png', as_attachment=True, download_name='qrcode.png')
    
    flash('No data provided for QR code.', 'error')
    return redirect(url_for('generate_qr'))

@app.route('/analytics')
@login_required
def analytics():
    """Analytics and reporting dashboard"""
    stats = get_analytics_stats()
    demographics = get_demographic_stats()
    event_analytics = get_event_analytics()
    
    return render_template('analytics.html', 
                         stats=stats, 
                         demographics=demographics,
                         event_analytics=event_analytics)

@app.route('/analytics/demographics')
@login_required
def demographics_report():
    """Detailed demographics report"""
    demographics = get_demographic_stats()
    return render_template('demographics_report.html', demographics=demographics)

@app.route('/analytics/events')
@login_required
def events_analytics():
    """Event analytics and trends"""
    event_analytics = get_event_analytics()
    return render_template('events_analytics.html', analytics=event_analytics)

@app.route('/search', methods=['GET', 'POST'])
@login_required
def advanced_search():
    """Advanced search across users, events, and attendance"""
    query = request.args.get('q', '') or request.form.get('q', '')
    search_type = request.args.get('type', 'all') or request.form.get('type', 'all')
    
    results = {
        'users': [],
        'events': [],
        'attendance': []
    }
    
    if query:
        conn = get_db()
        cursor = conn.cursor()
        
        if search_type in ['all', 'users']:
            # Search users
            cursor.execute('''
                SELECT * FROM users 
                WHERE name LIKE ? OR email LIKE ? OR id LIKE ? OR phone LIKE ? OR zone LIKE ?
                ORDER BY registration_date DESC
                LIMIT 50
            ''', (f'%{query}%', f'%{query}%', f'%{query}%', f'%{query}%', f'%{query}%'))
            results['users'] = [dict(row) for row in cursor.fetchall()]
        
        if search_type in ['all', 'events']:
            # Search events
            cursor.execute('''
                SELECT * FROM events 
                WHERE event_name LIKE ? OR event_description LIKE ? OR event_category LIKE ?
                ORDER BY event_date DESC
                LIMIT 50
            ''', (f'%{query}%', f'%{query}%', f'%{query}%'))
            results['events'] = [dict(row) for row in cursor.fetchall()]
        
        if search_type in ['all', 'attendance']:
            # Search attendance with user and event details
            cursor.execute('''
                SELECT a.*, u.name as user_name, u.email as user_email, u.id as user_id_display,
                       e.event_name, e.event_date
                FROM attendance a
                LEFT JOIN users u ON a.user_id = u.user_id
                LEFT JOIN events e ON a.event_id = e.event_id
                WHERE u.name LIKE ? OR u.email LIKE ? OR e.event_name LIKE ?
                ORDER BY a.attendance_date DESC
                LIMIT 50
            ''', (f'%{query}%', f'%{query}%', f'%{query}%'))
            results['attendance'] = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
    
    return render_template('advanced_search.html', results=results, query=query, search_type=search_type)

@app.route('/bulk-messaging', methods=['GET', 'POST'])
@login_required
def bulk_messaging():
    """Bulk messaging to users"""
    if request.method == 'POST':
        message_type = request.form.get('message_type', 'announcement')
        subject = request.form.get('subject', '').strip()
        message = request.form.get('message', '').strip()
        recipient_filter = request.form.get('recipient_filter', 'all')
        filter_value = request.form.get('filter_value', '').strip()
        
        if not subject or not message:
            flash('Please provide both subject and message.', 'error')
            return redirect(url_for('bulk_messaging'))
        
        # Get recipients based on filter
        conn = get_db()
        cursor = conn.cursor()
        
        if recipient_filter == 'all':
            cursor.execute('SELECT email, name FROM users WHERE email IS NOT NULL AND email != ""')
        elif recipient_filter == 'zone':
            cursor.execute('SELECT email, name FROM users WHERE zone = ? AND email IS NOT NULL AND email != ""', (filter_value,))
        elif recipient_filter == 'age_group':
            cursor.execute('SELECT email, name FROM users WHERE youth_age_group = ? AND email IS NOT NULL AND email != ""', (filter_value,))
        elif recipient_filter == 'classification':
            cursor.execute('SELECT email, name FROM users WHERE youth_classification = ? AND email IS NOT NULL AND email != ""', (filter_value,))
        else:
            cursor.execute('SELECT email, name FROM users WHERE email IS NOT NULL AND email != ""')
        
        recipients = cursor.fetchall()
        conn.close()
        
        # Send emails
        sent_count = 0
        failed_count = 0
        
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <h2 style="color: #002e6a;">{subject}</h2>
            <div style="white-space: pre-wrap;">{message}</div>
            <p style="margin-top: 2rem;">Best regards,<br>SAN AGUSTIN YEP ID System</p>
        </body>
        </html>
        """
        
        for recipient in recipients:
            email, name = recipient
            try:
                send_email_notification(
                    email,
                    subject,
                    message,
                    html_body.replace('{name}', name)
                )
                sent_count += 1
            except Exception as e:
                print(f"Error sending to {email}: {e}")
                failed_count += 1
        
        flash(f'Bulk message sent! {sent_count} emails sent successfully, {failed_count} failed.', 'success' if failed_count == 0 else 'warning')
        return redirect(url_for('bulk_messaging'))
    
    # Get filter options
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT DISTINCT zone FROM users WHERE zone IS NOT NULL AND zone != "" ORDER BY zone')
    zones = [row[0] for row in cursor.fetchall()]
    
    cursor.execute('SELECT DISTINCT youth_age_group FROM users WHERE youth_age_group IS NOT NULL AND youth_age_group != "" ORDER BY youth_age_group')
    age_groups = [row[0] for row in cursor.fetchall()]
    
    cursor.execute('SELECT DISTINCT youth_classification FROM users WHERE youth_classification IS NOT NULL AND youth_classification != "" ORDER BY youth_classification')
    classifications = [row[0] for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('bulk_messaging.html', zones=zones, age_groups=age_groups, classifications=classifications)

@app.route('/events/<event_id>/send-reminders', methods=['POST'])
@login_required
def send_event_reminders(event_id):
    """Send reminder emails for an event"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get event
    cursor.execute('SELECT * FROM events WHERE event_id = ?', (event_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        flash('Event not found.', 'error')
        return redirect(url_for('events'))
    event = dict(row)
    
    # Get all registered users
    cursor.execute('SELECT email, name FROM users WHERE email IS NOT NULL AND email != ""')
    users = cursor.fetchall()
    conn.close()
    
    sent_count = 0
    failed_count = 0
    
    for user_email, user_name in users:
        try:
            send_event_reminder(
                user_email,
                user_name,
                event.get('event_name', 'Event'),
                event.get('event_date', ''),
                event.get('event_time', '')
            )
            sent_count += 1
        except Exception as e:
            print(f"Error sending reminder to {user_email}: {e}")
            failed_count += 1
    
    # Mark reminders as sent
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE events SET reminder_sent = 1 WHERE event_id = ?', (event_id,))
    conn.commit()
    conn.close()
    
    flash(f'Event reminders sent! {sent_count} emails sent successfully, {failed_count} failed.', 'success' if failed_count == 0 else 'warning')
    return redirect(url_for('event_detail', event_id=event_id))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

